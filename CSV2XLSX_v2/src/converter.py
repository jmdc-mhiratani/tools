"""
CSV2XLSX 変換エンジン
高性能なCSV→Excel変換機能を提供
"""

from typing import Optional, Callable, Dict, Any
import pandas as pd
import chardet
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class CSVConverter:
    """CSV→Excel変換エンジン"""

    def __init__(self):
        self.encoding: Optional[str] = None
        self.delimiter: str = ','
        self.has_header: bool = True
        self.chunk_size: int = 10000  # 大容量ファイル対応

    def detect_encoding(self, file_path: Path) -> str:
        """
        ファイルのエンコーディングを検出

        Args:
            file_path: 対象ファイルパス

        Returns:
            検出されたエンコーディング名
        """
        try:
            with open(file_path, 'rb') as f:
                # ファイル全体を読み込んで精度を上げる
                raw_data = f.read()
                result = chardet.detect(raw_data)

            encoding = result.get('encoding', 'utf-8')
            confidence = result.get('confidence', 0.0)

            logger.info(f"Encoding detected: {encoding} (confidence: {confidence:.2f})")

            # 信頼度が低い場合のフォールバック
            if confidence < 0.7:
                # 日本語ファイルの一般的なエンコーディングを試行
                for fallback_encoding in ['utf-8', 'shift_jis', 'cp932']:
                    if self._test_encoding(file_path, fallback_encoding):
                        encoding = fallback_encoding
                        logger.warning(f"Low confidence, using fallback: {encoding}")
                        break

            return encoding

        except Exception as e:
            logger.error(f"Encoding detection failed: {e}")
            return 'utf-8'  # デフォルト

    def _test_encoding(self, file_path: Path, encoding: str) -> bool:
        """エンコーディングのテスト読み込み"""
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1000)  # 最初の1000文字を試行
            return True
        except (UnicodeDecodeError, UnicodeError):
            return False

    def detect_delimiter(self, file_path: Path, encoding: str) -> str:
        """
        CSVの区切り文字を自動検出

        Args:
            file_path: CSVファイルパス
            encoding: ファイルエンコーディング

        Returns:
            検出された区切り文字
        """
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                # 最初の数行を読み取り
                sample = f.read(1024)

            # 一般的な区切り文字をテスト
            delimiters = [',', '\t', ';', '|']
            delimiter_counts = {}

            for delimiter in delimiters:
                count = sample.count(delimiter)
                if count > 0:
                    delimiter_counts[delimiter] = count

            if delimiter_counts:
                # 最も多く使われている区切り文字を選択
                detected_delimiter = max(delimiter_counts, key=delimiter_counts.get)
                logger.info(f"Delimiter detected: '{detected_delimiter}'")
                return detected_delimiter

        except Exception as e:
            logger.error(f"Delimiter detection failed: {e}")

        return ','  # デフォルト

    def convert_to_excel(
        self,
        csv_path: Path,
        excel_path: Path,
        progress_callback: Optional[Callable[[int], None]] = None,
        style_options: Optional[Dict[str, Any]] = None
    ) -> bool:
        """
        CSVファイルをExcelに変換

        Args:
            csv_path: 入力CSVファイルパス
            excel_path: 出力Excelファイルパス
            progress_callback: 進捗コールバック関数
            style_options: スタイル設定オプション

        Returns:
            変換成功可否
        """
        try:
            logger.info(f"Converting {csv_path} to {excel_path}")

            # エンコーディングと区切り文字の検出
            self.encoding = self.detect_encoding(csv_path)
            self.delimiter = self.detect_delimiter(csv_path, self.encoding)

            if progress_callback:
                progress_callback(10)

            # ファイルサイズをチェックして処理方法を決定
            file_size = csv_path.stat().st_size
            if file_size > 50 * 1024 * 1024:  # 50MB以上
                return self._convert_large_file(csv_path, excel_path, progress_callback, style_options)
            else:
                return self._convert_standard_file(csv_path, excel_path, progress_callback, style_options)

        except Exception as e:
            logger.error(f"Conversion failed: {e}")
            return False

    def _convert_standard_file(
        self,
        csv_path: Path,
        excel_path: Path,
        progress_callback: Optional[Callable[[int], None]] = None,
        style_options: Optional[Dict[str, Any]] = None
    ) -> bool:
        """標準サイズファイルの変換"""
        try:
            # CSV読み込み
            df = pd.read_csv(
                csv_path,
                encoding=self.encoding,
                sep=self.delimiter,
                dtype=str  # データ型を保持
            )

            if progress_callback:
                progress_callback(50)

            # データ型の自動推定
            df = self._infer_data_types(df)

            if progress_callback:
                progress_callback(70)

            # Excel書き込み
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

                # スタイル適用
                if style_options:
                    worksheet = writer.sheets['Sheet1']
                    self._apply_styles(worksheet, df, style_options)

            if progress_callback:
                progress_callback(100)

            logger.info(f"Successfully converted {len(df)} rows to Excel")
            return True

        except Exception as e:
            logger.error(f"Standard file conversion failed: {e}")
            return False

    def _convert_large_file(
        self,
        csv_path: Path,
        excel_path: Path,
        progress_callback: Optional[Callable[[int], None]] = None,
        style_options: Optional[Dict[str, Any]] = None
    ) -> bool:
        """大容量ファイルのチャンク処理変換"""
        try:
            logger.info("Processing large file with chunking")

            # Excelワークブック作成
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sheet1'

            row_num = 1
            total_rows = 0

            # チャンク単位で処理
            for chunk_num, chunk in enumerate(pd.read_csv(
                csv_path,
                encoding=self.encoding,
                sep=self.delimiter,
                dtype=str,
                chunksize=self.chunk_size
            )):
                # データ型推定
                chunk = self._infer_data_types(chunk)

                # 最初のチャンクでヘッダーを追加
                if chunk_num == 0:
                    # ヘッダー行
                    for col_num, column_name in enumerate(chunk.columns, 1):
                        worksheet.cell(row=row_num, column=col_num, value=column_name)
                    row_num += 1

                # データ行
                for _, row in chunk.iterrows():
                    for col_num, value in enumerate(row, 1):
                        worksheet.cell(row=row_num, column=col_num, value=value)
                    row_num += 1

                total_rows += len(chunk)

                # 進捗更新
                if progress_callback:
                    progress = min(90, (chunk_num + 1) * 10)  # 最大90%まで
                    progress_callback(progress)

            # スタイル適用
            if style_options:
                self._apply_styles(worksheet, None, style_options)

            # ファイル保存
            workbook.save(excel_path)

            if progress_callback:
                progress_callback(100)

            logger.info(f"Successfully converted {total_rows} rows (large file)")
            return True

        except Exception as e:
            logger.error(f"Large file conversion failed: {e}")
            return False

    def _infer_data_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """データ型の自動推定"""
        try:
            # 数値型の推定
            for column in df.columns:
                # 数値変換を試行
                try:
                    # まず整数を試行
                    numeric_series = pd.to_numeric(df[column], errors='coerce')
                    if not numeric_series.isna().all():
                        # NaNでない値が存在する場合
                        if (numeric_series % 1 == 0).all():  # 全て整数
                            df[column] = numeric_series.astype('Int64')
                        else:  # 小数点を含む
                            df[column] = numeric_series
                except:
                    pass

                # 日付型の推定（数値列でない場合のみ）
                try:
                    # 既に数値型として処理された場合はスキップ
                    if pd.api.types.is_numeric_dtype(df[column]):
                        continue

                    date_series = pd.to_datetime(df[column], errors='coerce')
                    if not date_series.isna().all():
                        # 有効な日付が80%以上の場合に日付型として扱う
                        valid_dates = date_series.notna().sum()
                        if valid_dates / len(df) > 0.8:
                            df[column] = date_series
                except:
                    pass

        except Exception as e:
            logger.warning(f"Data type inference failed: {e}")

        return df

    def _apply_styles(
        self,
        worksheet,
        df: Optional[pd.DataFrame] = None,
        style_options: Dict[str, Any] = None
    ):
        """Excelスタイルの適用"""
        try:
            if not style_options:
                style_options = {}

            # ヘッダー行のスタイル
            header_font = Font(
                bold=style_options.get('header_bold', True),
                color=style_options.get('header_color', '000000')
            )
            header_fill = PatternFill(
                start_color=style_options.get('header_bg', 'E0E0E0'),
                end_color=style_options.get('header_bg', 'E0E0E0'),
                fill_type='solid'
            )

            # ヘッダー行にスタイル適用
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # 列幅の自動調整
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # 最大50文字
                worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.warning(f"Style application failed: {e}")


class ExcelToCSVConverter:
    """Excel→CSV変換エンジン"""

    def __init__(self):
        self.output_encoding = 'utf-8-sig'  # Excel互換のBOM付きUTF-8

    def convert_to_csv(
        self,
        excel_path: Path,
        csv_path: Path,
        sheet_name: Optional[str] = None,
        progress_callback: Optional[Callable[[int], None]] = None
    ) -> bool:
        """
        ExcelファイルをCSVに変換

        Args:
            excel_path: 入力Excelファイルパス
            csv_path: 出力CSVファイルパス
            sheet_name: 変換するシート名（Noneの場合は最初のシート）
            progress_callback: 進捗コールバック関数

        Returns:
            変換成功可否
        """
        try:
            logger.info(f"Converting {excel_path} to {csv_path}")

            if progress_callback:
                progress_callback(10)

            # Excel読み込み
            try:
                df = pd.read_excel(
                    excel_path,
                    sheet_name=sheet_name,
                    engine='openpyxl'
                )

                # 複数シートが返された場合は最初のシートを使用
                if isinstance(df, dict):
                    sheet_names = list(df.keys())
                    if sheet_names:
                        df = df[sheet_names[0]]
                        logger.info(f"Multiple sheets found, using: {sheet_names[0]}")
                    else:
                        logger.error("No sheets found in Excel file")
                        return False

            except Exception as e:
                logger.error(f"Excel file reading failed: {e}")
                return False

            if progress_callback:
                progress_callback(70)

            # CSV出力
            df.to_csv(
                csv_path,
                index=False,
                encoding=self.output_encoding
            )

            if progress_callback:
                progress_callback(100)

            logger.info(f"Successfully converted {len(df)} rows to CSV")
            return True

        except Exception as e:
            logger.error(f"Excel to CSV conversion failed: {e}")
            return False