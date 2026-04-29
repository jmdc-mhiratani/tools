"""
Excelスタイル適用モジュール
ワークシートへのスタイル、罫線、列幅調整などを適用
"""

import logging
from typing import Any, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.filters import AutoFilter
import pandas as pd

logger = logging.getLogger(__name__)


def apply_styles(
    worksheet,
    df: Optional[pd.DataFrame] = None,
    style_options: Optional[dict[str, Any]] = None,
    is_large_file: bool = False,
):
    """
    Excelスタイルの適用

    Args:
        worksheet: openpyxlワークシート
        df: データフレーム（日付列検出用）
        style_options: スタイルオプション辞書
        is_large_file: 大容量ファイルモード（スタイルを簡略化）
    """
    try:
        if is_large_file:
            logger.info(
                "Large file mode: applying simplified styles (header + auto-filter only)"
            )
            # 大容量ファイルではヘッダースタイルとオートフィルターのみ適用
            _apply_header_style(worksheet, style_options)
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            logger.info("Simplified styles applied successfully")
            return

        if not style_options:
            return

        # スタイルオプションをチェック
        apply_header_style = style_options.get(
            "header_bold", False
        ) or style_options.get("borders", False)
        auto_width = style_options.get("auto_width", False)
        freeze_header = style_options.get("freeze_header", False)
        alternating_rows = style_options.get("alternating_rows", False)

        # ヘッダー行のスタイル（header_bold or bordersが有効な場合）
        if apply_header_style:
            _apply_header_style(worksheet, style_options)

        # 罫線の設定
        if style_options.get("borders", False):
            _apply_borders(worksheet)

        # 交互の行色設定
        if alternating_rows and worksheet.max_row > 1:
            _apply_alternating_rows(worksheet)

        # ヘッダー固定（フリーズペイン）
        if freeze_header:
            worksheet.freeze_panes = "A2"

        # オートフィルターの設定
        if worksheet.max_row > 1 and worksheet.max_column > 0:
            filter_range = (
                f"A1:{worksheet.cell(row=1, column=worksheet.max_column).coordinate}"
            )
            worksheet.auto_filter = AutoFilter(ref=filter_range)

        # 日付列のフォーマット設定
        format_date_columns(worksheet, df)

        # 列幅の自動調整
        if auto_width:
            adjust_column_widths(worksheet)

    except Exception as e:
        logger.warning(f"Style application failed: {e}")


def _apply_header_style(worksheet, style_options: dict[str, Any]):
    """ヘッダー行のスタイル適用"""
    header_font = Font(
        bold=style_options.get("header_bold", True),
        color=style_options.get("header_color", "000000"),
    )
    header_fill = PatternFill(
        start_color=style_options.get("header_bg", "E0E0E0"),
        end_color=style_options.get("header_bg", "E0E0E0"),
        fill_type="solid",
    )

    # ヘッダー行のアライメント設定（文字折り返しを含む）
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,  # 文字の折り返し
    )

    # ヘッダー行にスタイル適用
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ヘッダー行の高さを55pxに設定（約41ポイント）
    worksheet.row_dimensions[1].height = 41.25  # 55px ≈ 41.25pt


def _apply_borders(worksheet):
    """罫線の適用"""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in worksheet.iter_rows(
        min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column
    ):
        for cell in row:
            cell.border = thin_border


def _apply_alternating_rows(worksheet):
    """交互の行色設定"""
    alternating_fill = PatternFill(
        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
    )
    for row_idx in range(2, worksheet.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if (
                    not cell.fill.start_color.rgb
                    or cell.fill.start_color.rgb == "00000000"
                ):
                    cell.fill = alternating_fill


def format_date_columns(worksheet, df: Optional[pd.DataFrame] = None):
    """
    日付列のフォーマット設定

    Args:
        worksheet: openpyxlワークシート
        df: データフレーム（日付列検出用）
    """
    try:
        if df is None:
            return

        # 日付列を検出してフォーマットを適用
        for col_idx, column in enumerate(df.columns, 1):
            if pd.api.types.is_datetime64_any_dtype(df[column]):
                # 日付列にフォーマット適用
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{column_letter}{row_idx}"]
                    if cell.value:
                        cell.number_format = "YYYY/M/D"

    except Exception as e:
        logger.warning(f"Date formatting failed: {e}")


def adjust_column_widths(worksheet):
    """
    列幅の自動調整

    Args:
        worksheet: openpyxlワークシート
    """
    try:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass

            # 最小12、最大50文字幅
            adjusted_width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        logger.warning(f"Column width adjustment failed: {e}")
